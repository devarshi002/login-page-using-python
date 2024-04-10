from tkinter import *
from tkinter.ttk import Combobox
from tkinter import Tk, messagebox
from tkcalendar import DateEntry
from babel import numbers

import openpyxl
import pathlib
from openpyxl import Workbook
import os

root = Tk()
root.title("Data Entry")
root.geometry('900x800+300+200')
root.resizable(True, True)
root.configure(bg="black")

file_path = 'patient_data.xlsx'
if os.path.exists(file_path):
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "PhoneNumber"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Occupation"
    sheet['F1'] = "Address"
    sheet['G1'] = "Cause"
    sheet['H1'] = "Past History"
    sheet['I1'] = "Family History"
    sheet['J1'] = "Treatment Date"
    sheet['K1'] = "Treatment"
    sheet['L1'] = "Fees"
    file.save('patient_data.xlsx')

def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    occupation = occupationValue.get()
    address = addressEntry.get(1.0, END)
    cause = causeValue.get()
    pastHistory = pastHistoryValue.get()
    familyHistory = familyHistoryValue.get()
    treatmentDate = treatmentDateEntry.get()
    treatmentFees = treatmentFeesValue.get()
    fees = feesValue.get()

    if not age.isdigit():
        messagebox.showerror('Error', 'Age must contain only digits.')
        return

    if len(age) > 3:
        messagebox.showerror('Error', 'Age must be up to 3 digits.')
        return

    if len(contact) != 10 or not contact.isdigit():
        messagebox.showerror('Error', 'Contact number must be a 10-digit number.')
        return

    file = openpyxl.load_workbook('patient_data.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=occupation)
    sheet.cell(column=6, row=sheet.max_row, value=address)
    sheet.cell(column=7, row=sheet.max_row, value=cause)
    sheet.cell(column=8, row=sheet.max_row, value=pastHistory)
    sheet.cell(column=9, row=sheet.max_row, value=familyHistory)
    sheet.cell(column=10, row=sheet.max_row, value=treatmentDate)
    sheet.cell(column=11, row=sheet.max_row, value=treatmentFees)
    sheet.cell(column=12, row=sheet.max_row, value=fees)
    file.save(r'patient_data.xlsx')
    messagebox.showinfo('info', 'detail added!')

    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    occupationValue.set('')
    addressEntry.delete(1.0, END)
    causeValue.set('')
    pastHistoryValue.set('')
    familyHistoryValue.set('')
    treatmentDateValue.set('')
    treatmentFeesValue.set('')
    feesValue.set('')

def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    occupationValue.set('')
    addressEntry.delete(1.0, END)
    causeValue.set('')
    pastHistoryValue.set('')
    familyHistoryValue.set('')
    treatmentDateValue.set('')
    treatmentFeesValue.set('')
    feesValue.set('')

def display_selected_date(event):
    selected_date = treatmentDateEntry.get_date()
    selected_date_label.config(text=f"Selected Date: {selected_date}")

def on_enter(event):
    event.widget.tk_focusNext().focus()
    return "break"

Label(root, text="PATIENTS DATA ENTRY:", font="arial 20", bg="green", fg="#fff").grid(row=0, column=0, columnspan=2, sticky="ew")

labels = ['Name', 'Contact No', 'Age', 'Gender', 'Occupation', 'Address', 'Cause', 'Past History', 'Family History', 'Treatment Date', 'Treatment', 'Fees']
for idx, label_text in enumerate(labels):
    Label(root, text=label_text, font='arial 12', bg='green', fg='#fff').grid(row=idx+1, column=0, padx=10, pady=5, sticky='w')

nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()
occupationValue = StringVar()
causeValue = StringVar()
pastHistoryValue = StringVar()
familyHistoryValue = StringVar()
treatmentDateValue = StringVar()
treatmentFeesValue = StringVar()
feesValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=20, bd=2, font=20)
ageEntry = Spinbox(root, from_=0, to=150, textvariable=AgeValue, width=15, bd=2, font=20)
occupationEntry = Entry(root, textvariable=occupationValue, width=30, bd=2, font=20)
causeEntry = Entry(root, textvariable=causeValue, width=45, bd=2, font=20)
pastHistoryEntry = Entry(root, textvariable=pastHistoryValue, width=45, bd=2, font=20)
familyHistoryEntry = Entry(root, textvariable=familyHistoryValue, width=45, bd=2, font=20)
treatmentFeesEntry = Entry(root, textvariable=treatmentFeesValue, width=45, bd=2, font=20)
feesEntry = Entry(root, textvariable=feesValue, width=10, bd=2, font=20)
gender_combobox = Combobox(root, values=['Male', 'Female', 'Other'], font='arial 14', state='r', width=14)
gender_combobox.grid(row=4, column=1, padx=10, pady=5, sticky='w')
gender_combobox.set('Male')

addressEntry = Text(root, width=50, height=2, bd=2)
nameEntry.grid(row=1, column=1, padx=10, pady=5, sticky='ew')
contactEntry.grid(row=2, column=1, padx=10, pady=5, sticky='ew')
ageEntry.grid(row=3, column=1, padx=10, pady=5, sticky='ew')
occupationEntry.grid(row=5, column=1, padx=10, pady=5, sticky='ew')
causeEntry.grid(row=7, column=1, padx=10, pady=5, sticky='ew')
pastHistoryEntry.grid(row=8, column=1, padx=10, pady=5, sticky='ew')
familyHistoryEntry.grid(row=9, column=1, padx=10, pady=5, sticky='ew')
treatmentFeesEntry.grid(row=11, column=1, padx=10, pady=5, sticky='ew')
addressEntry.grid(row=6, column=1, padx=10, pady=5, sticky='ew')
feesEntry.grid(row=12, column=1, padx=10, pady=5, sticky='ew')

treatmentDateEntry = DateEntry(root, width=30, bd=2, font=20)
treatmentDateEntry.grid(row=10, column=1, padx=10, pady=5, sticky='ew')
treatmentDateEntry.bind("<<DateEntrySelected>>", display_selected_date)

selected_date_label = Label(root, text="Selected Date:", font='arial 12', bg='green', fg='#fff')
selected_date_label.grid(row=10, column=1, padx=10, pady=5, sticky='w')

Button(root, text="Submit", bg="#326273", fg="white", width=5, height=2, command=submit).grid(row=13, column=0, padx=10, pady=10, sticky='ew')
Button(root, text="Clear", bg="#326273", fg="white", width=5, height=2, command=clear).grid(row=13, column=1, padx=10, pady=10, sticky='ew')
Button(root, text="Exit", bg="#326273", fg="white", width=5, height=2, command=root.destroy).grid(row=13, column=2, padx=10, pady=10, sticky='ew')

# Configure rows and columns to resize proportionally
for i in range(len(labels) + 2):  # +2 for the title row and button row
    root.grid_rowconfigure(i, weight=1)
    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)
    

# Binding Enter key press event to move focus to next widget
nameEntry.bind('<Return>', on_enter)
contactEntry.bind('<Return>', on_enter)
ageEntry.bind('<Return>', on_enter)
occupationEntry.bind('<Return>', on_enter)
causeEntry.bind('<Return>', on_enter)
pastHistoryEntry.bind('<Return>', on_enter)
familyHistoryEntry.bind('<Return>', on_enter)
treatmentDateEntry.bind('<Return>', on_enter)
treatmentFeesEntry.bind('<Return>', on_enter)

root.mainloop()
